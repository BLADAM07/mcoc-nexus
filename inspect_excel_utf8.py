import openpyxl
import codecs

file_path = 'nexus-frontend/public/excle/MCOC_dataset.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True)

with codecs.open('excel_preview_utf8.txt', 'w', encoding='utf-8') as f:
    f.write("Sheets: " + str(wb.sheetnames) + "\n")

    for sheet_name in wb.sheetnames:
        if "Story" in sheet_name or "Act" in sheet_name or "Node" in sheet_name or "Story_nodes" in sheet_name:
            f.write(f"\n--- {sheet_name} ---\n")
            sheet = wb[sheet_name]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                f.write(str(row) + "\n")
                if i > 5:
                    break
