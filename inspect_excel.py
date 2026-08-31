import openpyxl

file_path = 'nexus-frontend/public/excle/MCOC_dataset.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    if "Story" in sheet_name or "Act" in sheet_name or "Node" in sheet_name:
        print(f"\n--- {sheet_name} ---")
        sheet = wb[sheet_name]
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            print(row)
            if i > 5:
                break
